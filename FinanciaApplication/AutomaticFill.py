# -*- coding: utf-8 -*-
"""
Created on Sat Oct 31 00:25:36 2020

@author: alexandru.vesa
"""
import pandas as pd
from openpyxl.utils import FORMULAE  
from openpyxl import load_workbook , Workbook
import xlwings as xw
pathOrig  = r'C:\Users\alexandru.vesa\Desktop\Research\New_Personal_Program_DL_Programming\GitMy\my-work\FinanciaApplication\golOrig.xlsx'
pathSim = r'C:\Users\alexandru.vesa\Desktop\Research\New_Personal_Program_DL_Programming\GitMy\my-work\FinanciaApplication\golOrig.xlsx'



""""
Pt proiectul cu investitie acolo unde ar trebui sa introducem manual datele ne putem
lua pt primul an dupa proiectul fara investitie , apoi facem ajustari automate
in functie de IRR
"""

def modificareTabel1(pathOrig, pathSim):
    

    workbook = load_workbook(filename = pathOrig)
    workbookSim=load_workbook(filename =pathSim)
    
    ProiectiiFinInvestitie = workbook['3A-Proiectii_fin_investitie']
    #Bilant = workbook['1A-Bilant']
    #PP = workbook['1B-ContPP']
    
    
    ProiectiiInvestitieSim=workbook['3A-Proiectii_fin_investitie']
    
    
    #Modify the columns from ProiectiiFinInvestitie based on Bilant and PP
    
    """
    INCASARI DIN ACTIVITATEA DE EXPLOATARE (fara investitie)
    """
    ###
    PretUnitarMarfaOrig = ProiectiiFinInvestitie['D17'].value 
    ProiectiiInvestitieSim['D17'] = PretUnitarMarfaOrig
    ProiectiiInvestitieSim['E17']= '=D17*1.01'#pret unitar (marfa)
    ProiectiiInvestitieSim['F17'] ='=E17*1.01'#pret unitar (marfa)
    ProiectiiInvestitieSim['G17'] ='=F17*1.01' #pret unitar (marfa)
    
    ProiectiiInvestitieSim['D15'] = '=D16*D17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['E15'] = '=E16*E17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['F15'] = '=F16*F17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['G15'] = '=G16*G17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['C15'] = '=SUM(D15:G15)' #Total
    
    ProiectiiInvestitieSim['D18'] = '=D9+D12+D15'
    ProiectiiInvestitieSim['E18'] = '=E9+E12+E15'
    ProiectiiInvestitieSim['F18'] = '=F9+F12+F15'
    ProiectiiInvestitieSim['G18'] = '=G9+G12+G15'
    ProiectiiInvestitieSim['C18'] = '=SUM(D18:G18)'#Total incasari 
    
    ProiectiiInvestitieSim['D16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['E16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['F16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['G16'] =1 #cantitate marfuri
    ###
    
    """
    Cheltuieli de exploatare
    """
    ### 5.Cheltuieli cu materiile prime si cu materialele consumabile
    PretUnitarMateriiPrimeOrig = ProiectiiFinInvestitie['D23'].value
    ProiectiiInvestitieSim['D23'] = PretUnitarMateriiPrimeOrig
    ProiectiiInvestitieSim['E23'] = '=D23*1.01'
    ProiectiiInvestitieSim['F23'] = '=E23*1.01'
    ProiectiiInvestitieSim['G23'] = '=F23*1.01'
    
    ProiectiiInvestitieSim['D22'] = 1
    ProiectiiInvestitieSim['E22'] = 1
    ProiectiiInvestitieSim['F22'] = 1
    ProiectiiInvestitieSim['G22'] = 1
    
    ProiectiiInvestitieSim['D24'] = 1
    ProiectiiInvestitieSim['E24'] = 1
    ProiectiiInvestitieSim['F24'] = 1
    ProiectiiInvestitieSim['G24'] = 1
    
    PretUnitarMaterialeConsumabile = ProiectiiFinInvestitie['D25'].value
    ProiectiiInvestitieSim['D25'] =PretUnitarMaterialeConsumabile
    ProiectiiInvestitieSim['E25'] ='=D25*1.05'
    ProiectiiInvestitieSim['F25'] ='=E25*1.05'
    ProiectiiInvestitieSim['G25'] ='=F25*1.05'
    
    ProiectiiInvestitieSim['D21'] ='=SUM(D22*D23)+SUM(D24*D25)'
    ProiectiiInvestitieSim['E21'] ='=SUM(E22*E23)+SUM(E24*E25)'
    ProiectiiInvestitieSim['F21'] ='=SUM(F22*F23)+SUM(F24*F25)'
    ProiectiiInvestitieSim['G21'] ='=SUM(G22*G23)+SUM(G24*G25)'
    ProiectiiInvestitieSim['C21'] = '=SUM(D21:G21)'
    
    ### 6.Cheltuieli privind marfurile 
    PretUnitarMarfuri = ProiectiiFinInvestitie['D28'].value
    ProiectiiInvestitieSim['D28'] = PretUnitarMarfuri
    ProiectiiInvestitieSim['E28'] = 0
    ProiectiiInvestitieSim['F28'] = 0
    ProiectiiInvestitieSim['G28'] = 0
    
    ProiectiiInvestitieSim['D27'] = 1
    ProiectiiInvestitieSim['E27'] = 0
    ProiectiiInvestitieSim['F27'] = 0
    ProiectiiInvestitieSim['G27'] = 0
    
    ProiectiiInvestitieSim['D26'] = '=D27*D28'
    ProiectiiInvestitieSim['E26'] = '=E27*E28'
    ProiectiiInvestitieSim['F26'] = '=F27*F28'
    ProiectiiInvestitieSim['G26'] = '=G27*G28'
    
    
    ### 7.Alte cheltuieli materiale (inclusiv cheltuieli cu prestatii externe)
    AlteCheltuieliMateriale = ProiectiiFinInvestitie['D29'].value
    ProiectiiInvestitieSim['D29'] = AlteCheltuieliMateriale
    ProiectiiInvestitieSim['E29'] = '=D29*1.01'
    ProiectiiInvestitieSim['F29'] = '=E29*1.01'
    ProiectiiInvestitieSim['G29'] = '=F29*1.01'
    ProiectiiInvestitieSim['C29']='=SUM(D29:G29)' #Total
    
    
    ### 8.Cheltuieli cu energia 
    ProiectiiInvestitieSim['D31'] = 1
    ProiectiiInvestitieSim['E31'] = 1
    ProiectiiInvestitieSim['F31'] = 1
    ProiectiiInvestitieSim['G31'] = 1
    
    TarifFurnizareUnitar = ProiectiiFinInvestitie['D32'].value
    ProiectiiInvestitieSim['D32'] = TarifFurnizareUnitar
    ProiectiiInvestitieSim['E32']='=D32*1.01'
    ProiectiiInvestitieSim['F32']='=E32*1.01'
    ProiectiiInvestitieSim['G32']='=F32*1.01'
    
    ProiectiiInvestitieSim['D30'] = '=D31*D32'
    ProiectiiInvestitieSim['E30'] = '=E31*E32'
    ProiectiiInvestitieSim['F30'] = '=F31*F32'
    ProiectiiInvestitieSim['G30'] = '=G31*G32'
    
    ProiectiiInvestitieSim['C30'] ='=SUM(D30:G30)' #Total 
    
    ### 9.Cheltuieli cu apa
    
    ### 10.Alte cheltuieli (utilitati)
    
    
    ### Total Cheltuieli Materiale
    ProiectiiInvestitieSim['D39']='=D21+D26+D29+D30+D33+D36'
    ProiectiiInvestitieSim['E39']='=E21+E26+E29+E30+E33+E36'
    ProiectiiInvestitieSim['F39']='=F21+F26+F29+F30+F33+F36'
    ProiectiiInvestitieSim['G39']='=G21+G26+G29+G30+G33+G36'
    
    ProiectiiInvestitieSim['C39']= '=SUM(D39:G39)'
    
    
    ### 11.Cheltuieli cu personalul angajat
    ProiectiiInvestitieSim['D41']=1
    ProiectiiInvestitieSim['E41']=1
    ProiectiiInvestitieSim['F41']=1
    ProiectiiInvestitieSim['G41']=1
    
    SalariuDeBaza =ProiectiiFinInvestitie['D42'].value
    ProiectiiInvestitieSim['D42'] = SalariuDeBaza
    ProiectiiInvestitieSim['E42']='=D42*1'
    ProiectiiInvestitieSim['F42']='=E42*1'
    ProiectiiInvestitieSim['G42']='=F42*1'
    
    ProiectiiInvestitieSim['D43']=1
    ProiectiiInvestitieSim['E43']=1
    ProiectiiInvestitieSim['F43']=1
    ProiectiiInvestitieSim['G43']=1
    
    ProiectiiInvestitieSim['D40']='=(D41*D42)*D43'
    ProiectiiInvestitieSim['E40']='=(E41*E42)*E43'
    ProiectiiInvestitieSim['F40']='=(F41*F42)*F43'
    ProiectiiInvestitieSim['G40']='=(G41*G42)*G43'
    
    ProiectiiInvestitieSim['C40']='=SUM(D40:G40)' #Total
    
    ### 12.Cheltuieli cu asigurarile si protectia sociala 
    ProiectiiInvestitieSim['D44']=0
    ProiectiiInvestitieSim['E44']=0
    ProiectiiInvestitieSim['F44']=0
    ProiectiiInvestitieSim['G44']=0
    ProiectiiInvestitieSim['C44']= '=SUM(D44:G44)' #Total
    
    ### Cheltuieli de personal 
    ProiectiiInvestitieSim['D45'] = '=D40+D44'
    ProiectiiInvestitieSim['E45'] = '=E40+E44'
    ProiectiiInvestitieSim['F45'] = '=F40+F44'
    ProiectiiInvestitieSim['G45'] = '=G40+G44'
    ProiectiiInvestitieSim['C45']= '=SUM(D45:G45)' #Total
    
    
    ### 13.Alte cheltuieli de exploatare (prestatii externe, alte impozite, taxe si varsaminte asimilate, alte cheltuieli), din care:
    Tarif = ProiectiiFinInvestitie['D49'].value
    ProiectiiInvestitieSim['D49'] = Tarif
    ProiectiiInvestitieSim['E49'] = '=D49*1.01'
    ProiectiiInvestitieSim['F49'] = '=E49*1.01'
    ProiectiiInvestitieSim['G49'] = '=F49*1.01'
    
    ProiectiiInvestitieSim['D48'] = 1
    ProiectiiInvestitieSim['E48'] = 1
    ProiectiiInvestitieSim['F48'] = 1
    ProiectiiInvestitieSim['G48'] = 1
    
    ProiectiiInvestitieSim['D47'] = '=D48*D49'
    ProiectiiInvestitieSim['E47'] = '=E48*E49'
    ProiectiiInvestitieSim['F47'] = '=F48*F49'
    ProiectiiInvestitieSim['G47'] = '=G48*G49'
    
    ProiectiiInvestitieSim['C47']='=SUM(D47:G47)' #Total
    
    Altele = ProiectiiFinInvestitie['D46'].value
    ProiectiiInvestitieSim['D46'] = Altele
    ProiectiiInvestitieSim['E46'] ='=D46*1.05'
    ProiectiiInvestitieSim['F46'] ='=E46*1.05'
    ProiectiiInvestitieSim['G46'] ='=F46*1.05'
    
    ProiectiiInvestitieSim['C46']='=SUM(D46:G46)' #Total
    
    
    
    
    ### 14.Cheltuieli financiare 
    CheltuieliFinanciare = ProiectiiFinInvestitie['D50'].value
    ProiectiiInvestitieSim['D50'] = CheltuieliFinanciare
    ProiectiiInvestitieSim['E50']='=D50*1.01'
    ProiectiiInvestitieSim['F50']='=E50*1.01'
    ProiectiiInvestitieSim['G50']='=F50*1.01'
    ProiectiiInvestitieSim['C50']='=SUM(D50:G50)'
    
    # Disponibil de numerar la inceputul perioadei 
    Numerar = ProiectiiFinInvestitie['D58'].value
    ProiectiiInvestitieSim['D58'] = Numerar
    
        
    ### INCASARI DIN ACTIVITATEA DE EXPLOATARE  (cu adoptarea investitiei)
    
    ### 2.Venituri din prestari servicii
    Venit = ProiectiiFinInvestitie['D70'].value
    ProiectiiInvestitieSim['D70'] = Venit
    ProiectiiInvestitieSim['E70'] = 620000
    ProiectiiInvestitieSim['F70'] = '=E70*1.02'
    ProiectiiInvestitieSim['G70'] = '=F70*1.03'

    ProiectiiInvestitieSim['D69'] = 1
    ProiectiiInvestitieSim['E69'] = 1
    ProiectiiInvestitieSim['F69'] = 1
    ProiectiiInvestitieSim['G69'] = 1
    
    ProiectiiInvestitieSim['D68']='=D69*D70'
    ProiectiiInvestitieSim['E68']='=E69*E70'
    ProiectiiInvestitieSim['F68']='=F69*F70'
    ProiectiiInvestitieSim['G68']='=G69*G70'
    
    ProiectiiInvestitieSim['C68']='=SUM(D68:G68)' #Total
    
    ### 3.Venituri din vanzari marfuri
    ProiectiiInvestitieSim['D73'] = '=D17'
    ProiectiiInvestitieSim['E73'] = '=D73*1.02'
    ProiectiiInvestitieSim['F73'] = '=E73*1.02'
    ProiectiiInvestitieSim['G73'] = '=F73*1.02'

    ProiectiiInvestitieSim['D72'] =1
    ProiectiiInvestitieSim['E72'] =1
    ProiectiiInvestitieSim['F72'] =1
    ProiectiiInvestitieSim['G72'] =1
    
    ProiectiiInvestitieSim['D71'] = '=D72*D73'
    ProiectiiInvestitieSim['E71'] = '=E72*E73'
    ProiectiiInvestitieSim['F71'] = '=F72*F73'
    ProiectiiInvestitieSim['G71'] = '=G72*G73'
    
    ProiectiiInvestitieSim['C71']= '=SUM(D71:G71)' #Total
    
    ProiectiiInvestitieSim['D74']= '=D65+D68+D71'
    ProiectiiInvestitieSim['E74']= '=E65+E68+E71'
    ProiectiiInvestitieSim['F74']= '=F65+F68+F71'
    ProiectiiInvestitieSim['G74']= '=G65+G68+G71'

    ProiectiiInvestitieSim['C74']= '=SUM(D74:G74)' #Total
    
    ### 5.Cheltuieli cu materiile prime si cu materialele consumabile
    
    ProiectiiInvestitieSim['D79'] =1008702
    ProiectiiInvestitieSim['E79'] = '=D79*1.02'
    ProiectiiInvestitieSim['F79'] = '=E79*1.03'
    ProiectiiInvestitieSim['G79'] = '=F79*1.03'

    ProiectiiInvestitieSim['D78'] = 1
    ProiectiiInvestitieSim['E78'] = 1
    ProiectiiInvestitieSim['F78'] = 1
    ProiectiiInvestitieSim['G78'] = 1
    
    
    ProiectiiInvestitieSim['D77'] = '=SUM(D78*D79)+SUM(D80*D81)'
    ProiectiiInvestitieSim['E77'] = '=SUM(E78*E79)+SUM(E80*E81)'
    ProiectiiInvestitieSim['F77'] = '=SUM(F78*F79)+SUM(F80*F81)'
    ProiectiiInvestitieSim['G77'] = '=SUM(G78*G79)+SUM(G80*G81)'

    ProiectiiInvestitieSim['C77']= '=SUM(D77:G77)' #Total
    
    
    
    
    ### 6.Cheltuieli privind marfurile 
    
    ProiectiiInvestitieSim['D84'] = 0
    ProiectiiInvestitieSim['E84'] = 0
    ProiectiiInvestitieSim['F84'] = 0
    ProiectiiInvestitieSim['G84'] = 0
    
    ProiectiiInvestitieSim['D83'] = 0
    ProiectiiInvestitieSim['E83'] = 0
    ProiectiiInvestitieSim['F83'] = 0
    ProiectiiInvestitieSim['G83'] = 0
    
    ProiectiiInvestitieSim['D82'] = '=D83*D84'
    ProiectiiInvestitieSim['E82'] = '=E83*E84'
    ProiectiiInvestitieSim['F82'] = '=F83*F84'
    ProiectiiInvestitieSim['G82'] = '=G83*G84'
    
    ProiectiiInvestitieSim['C82'] = '=SUM(D82:G82)' #Total
    
    
    ### 7.Alte cheltuieli materiale (inclusiv cheltuieli cu prestatii externe)
    
    #Alte cheltuieli materiale  fara invvestitie == Alte cheltuieli investitie 
    ProiectiiInvestitieSim['D85'] = ProiectiiInvestitieSim['D29'].value
    ProiectiiInvestitieSim['E85'] ='=D85*1.02'
    ProiectiiInvestitieSim['F85'] ='=E85*1.02'
    ProiectiiInvestitieSim['G85'] ='=F85*1.02'
    
    ProiectiiInvestitieSim['C85']='=SUM(D85:G85)' #Total
    
    ### 8.Cheltuieli cu energia +apa
    
    
    ### 9.Cheltuieli apa
    
    
    ### 10.Alte cheltuieli
    
    
    ### Total Cheltuieli Materiale
    ProiectiiInvestitieSim['D95']='=D77+D82+D85+D86+D89+D92'
    ProiectiiInvestitieSim['E95']='=E77+E82+E85+E86+E89+E92'
    ProiectiiInvestitieSim['F95']='=F77+F82+F85+F86+F89+F92'
    ProiectiiInvestitieSim['G95']='=G77+G82+G85+G86+G89+G92'
    
    ProiectiiInvestitieSim['C95']= '=SUM(D95:G95)'
    
    
    
    
    ### 11.Cheltuieli cu personal
    ProiectiiInvestitieSim['D99']=1
    ProiectiiInvestitieSim['E99']=1
    ProiectiiInvestitieSim['F99']=1
    ProiectiiInvestitieSim['G99']=1
    
    #Salar cu investitiii = Salar fara investitii pt primul an
    ProiectiiInvestitieSim['D98'] = ProiectiiInvestitieSim['D42'].value
    ProiectiiInvestitieSim['E98']='=D98*1.1'
    ProiectiiInvestitieSim['F98']='=E98*1.1'
    ProiectiiInvestitieSim['G98']='=F98*1.1'
    
    ProiectiiInvestitieSim['D97'] =1
    ProiectiiInvestitieSim['E97'] =1
    ProiectiiInvestitieSim['F97']=1
    ProiectiiInvestitieSim['G97'] =1
    
    ProiectiiInvestitieSim['D96'] = '=SUM(D97*D98)*D99'
    ProiectiiInvestitieSim['E96'] = '=SUM(E97*E98)*E99'
    ProiectiiInvestitieSim['F96'] = '=SUM(F97*F98)*F99'
    ProiectiiInvestitieSim['G96'] = '=SUM(G97*G98)*G99'

    ProiectiiInvestitieSim['C96'] ='=SUM(D96:G96)'

    # 12.Cheltuieli cu asigurarile si protectia sociala 
    ProiectiiInvestitieSim['D100'] = '=D96*2.25%'
    ProiectiiInvestitieSim['E100'] = '=E96*2.25%'
    ProiectiiInvestitieSim['F100'] = '=F96*2.25%'
    ProiectiiInvestitieSim['G100'] = '=G96*2.25%'
    
    ProiectiiInvestitieSim['D101'] = '=D100+D96'
    ProiectiiInvestitieSim['E101'] = '=E100+E96'
    ProiectiiInvestitieSim['F101'] = '=F100+F96'
    ProiectiiInvestitieSim['G101'] = '=G100+G96'

    ProiectiiInvestitieSim['C101'] ='=SUM(D101:G101)'

    # 13.Alte cheltuieli de exploatare (prestatii externe, alte impozite, taxe si varsaminte asimilate, alte cheltuieli), din care:
    AlteCheltuieli = ProiectiiFinInvestitie['D102'].value
    ProiectiiInvestitieSim['D102'] = AlteCheltuieli
    ProiectiiInvestitieSim['E102']='=D102*1.05'
    ProiectiiInvestitieSim['F102'] ='=E102*1.05'
    ProiectiiInvestitieSim['G102'] ='=F102*1.05'
    
    ProiectiiInvestitieSim['C102'] = ' =SUM(D102:G102)'

    
    ProiectiiInvestitieSim['D101'] ='=D100+D96'
    ProiectiiInvestitieSim['E101'] ='=E100+E96'
    ProiectiiInvestitieSim['F101'] = '=F100+F96'
    ProiectiiInvestitieSim['G101'] = '=G100+G96'

    ProiectiiInvestitieSim['C101'] = '=SUM(D101:G101)'
    
    
    #Tarif proiecti investitie = Tarif proiect fara investitie (*primul an )
    ProiectiiInvestitieSim['D105'] = ProiectiiInvestitieSim['D49'].value
    ProiectiiInvestitieSim['E105'] ='=D105*1.02'
    ProiectiiInvestitieSim['F105'] ='=E105*1.03'
    ProiectiiInvestitieSim['G105'] ='=F105*1.04'
    
    ProiectiiInvestitieSim['D104'] =1
    ProiectiiInvestitieSim['E104'] =1
    ProiectiiInvestitieSim['F104'] =1
    ProiectiiInvestitieSim['G104'] =1
    
    ProiectiiInvestitieSim['D103']='=D104*D105'
    ProiectiiInvestitieSim['E103']='=E104*E105'
    ProiectiiInvestitieSim['F103']='=F104*F105'
    ProiectiiInvestitieSim['G103']='=G104*G105'

    ProiectiiInvestitieSim['C103']='=SUM(D103:G103)'


    workbook.save(filename='fuck.xlsx')
    
    
  





    

modificareTabel1(pathOrig, pathSim)








