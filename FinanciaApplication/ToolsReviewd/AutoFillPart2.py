# -*- coding: utf-8 -*-
"""
Created on Fri Nov  6 19:00:26 2020

@author: alexandru.vesa
"""
import pandas as pd
from openpyxl.utils import FORMULAE  
from openpyxl import load_workbook , Workbook
import xlwings as xw
import os
import argparse




""""
Pt proiectul cu investitie acolo unde ar trebui sa introducem manual datele ne putem
lua pt primul an dupa proiectul fara investitie , apoi facem ajustari automate
in functie de IRR
"""

def modificareTabel1(path):
    

    #workbook = load_workbook(filename = pathOrig)
    workbookSim=load_workbook(filename =path)
    
    #ProiectiiFinInvestitie = workbook['3A-Proiectii_fin_investitie']
    #Bilant = workbook['1A-Bilant']
    #PP = workbook['1B-ContPP']
    
    
    ProiectiiInvestitieSim=workbookSim['3A-Proiectii_fin_investitie']
    ProiectiiRentabilitate = workbookSim['3B-Rentabilitate_investitie']
    
    
    ######################################################################## Sheet4
    ProiectiiFinaleInvestitie =  workbookSim['4-Proiectii_fin_intreprindere']
    
    
    ProiectiiFinaleInvestitie['C20'] = "='1B-ContPP'!C56 + 0.5*'1B-ContPP'!C56"
    
    ProiectiiFinaleInvestitie['C38'] = "='3A-Proiectii_fin_investitie'!D65/1.19"
    ProiectiiFinaleInvestitie['D38'] = "='3A-Proiectii_fin_investitie'!E65/1.19"
    ProiectiiFinaleInvestitie['E38'] = "='3A-Proiectii_fin_investitie'!F65/1.19"
    ProiectiiFinaleInvestitie['F38'] = "='3A-Proiectii_fin_investitie'!G65/1.19"

    
    ProiectiiFinaleInvestitie['C39'] = "=C38*0.19"
    ProiectiiFinaleInvestitie['D39'] = "=D38*0.19"
    ProiectiiFinaleInvestitie['E39'] = "=E38*0.19"
    ProiectiiFinaleInvestitie['F39'] = "=F38*0.19"

    
    ProiectiiFinaleInvestitie['C41'] =  "='3A-Proiectii_fin_investitie'!D68/1.19"
    ProiectiiFinaleInvestitie['D41'] =   "='3A-Proiectii_fin_investitie'!E68/1.19"
    ProiectiiFinaleInvestitie['E41'] =   "='3A-Proiectii_fin_investitie'!F68/1.19"
    ProiectiiFinaleInvestitie['F41'] =   "='3A-Proiectii_fin_investitie'!G68/1.19"

    
    ProiectiiFinaleInvestitie['C42'] = "=C41*0.19"
    ProiectiiFinaleInvestitie['E42'] = "=D41*0.19"
    ProiectiiFinaleInvestitie['D42'] = "=E41*0.19"
    ProiectiiFinaleInvestitie['F42'] = "=F41*0.19"

    ProiectiiFinaleInvestitie['C44'] = "='3A-Proiectii_fin_investitie'!D73/1.19"
    ProiectiiFinaleInvestitie['D44'] = "='3A-Proiectii_fin_investitie'!E73/1.19"
    ProiectiiFinaleInvestitie['E44'] = "='3A-Proiectii_fin_investitie'!F73/1.19"
    ProiectiiFinaleInvestitie['F44'] = "='3A-Proiectii_fin_investitie'!G73/1.19"

    ProiectiiFinaleInvestitie['C45'] = "=C44*0.19"
    ProiectiiFinaleInvestitie['D45'] = "=D44*0.19"
    ProiectiiFinaleInvestitie['E45'] = "=E44*0.19"
    ProiectiiFinaleInvestitie['F45'] = "=F44*0.19"

    ProiectiiFinaleInvestitie['C47']= "='1B-ContPP'!C11"
    
    ProiectiiFinaleInvestitie['C48']= "=C47*0.19"
    ProiectiiFinaleInvestitie['D48']= "=D47*0.19"
    ProiectiiFinaleInvestitie['E48']= "=E47*0.19"
    ProiectiiFinaleInvestitie['F48']= "=F47*0.19"

    ProiectiiFinaleInvestitie['C53'] = "='1B-ContPP'!C12 "
    
    ProiectiiFinaleInvestitie['C54'] = "=C53*0.19"
    ProiectiiFinaleInvestitie['D54'] = "=D53*0.19"
    ProiectiiFinaleInvestitie['E54'] = "=E53*0.19"
    ProiectiiFinaleInvestitie['F54'] = "=F53*0.19"

    
    ProiectiiFinaleInvestitie['C56'] = "='1B-ContPP'!C28"
    ProiectiiFinaleInvestitie['C57'] ="='1B-ContPP'!C29 "
    ProiectiiFinaleInvestitie['C58']="='1B-ContPP'!C51"
    
    ProiectiiFinaleInvestitie['C64'] = "='3A-Proiectii_fin_investitie'!D77/1.19"
    ProiectiiFinaleInvestitie['D64'] = "='3A-Proiectii_fin_investitie'!E77/1.19"
    ProiectiiFinaleInvestitie['E64'] = "='3A-Proiectii_fin_investitie'!F77/1.19"
    ProiectiiFinaleInvestitie['F64'] = "='3A-Proiectii_fin_investitie'!G77/1.19"

    ProiectiiFinaleInvestitie['C65'] ="=C64*0.19"
    ProiectiiFinaleInvestitie['D65'] ="=D64*0.19"
    ProiectiiFinaleInvestitie['E65'] ="=E64*0.19"
    ProiectiiFinaleInvestitie['F65'] ="=F64*0.19"

    
    
    ProiectiiFinaleInvestitie['C67'] = "='3A-Proiectii_fin_investitie'!D85 / 1.19"
    ProiectiiFinaleInvestitie['D67'] = "='3A-Proiectii_fin_investitie'!E85 / 1.19"
    ProiectiiFinaleInvestitie['E67'] = "='3A-Proiectii_fin_investitie'!F85 / 1.19"
    ProiectiiFinaleInvestitie['F67'] = "='3A-Proiectii_fin_investitie'!G85 / 1.19"

    ProiectiiFinaleInvestitie['C68'] = "=C67*0.19"
    ProiectiiFinaleInvestitie['D68'] = "=D67*0.19"
    ProiectiiFinaleInvestitie['E68'] = "=E67*0.19"
    ProiectiiFinaleInvestitie['F68'] = "=F67*0.19"

    
    
    ProiectiiFinaleInvestitie['C70'] = "=('3A-Proiectii_fin_investitie'!D86 +'3A-Proiectii_fin_investitie'!D89) /1.19 "
    ProiectiiFinaleInvestitie['D70'] = "=('3A-Proiectii_fin_investitie'!E86 +'3A-Proiectii_fin_investitie'!E89) /1.19 "
    ProiectiiFinaleInvestitie['E70'] = "=('3A-Proiectii_fin_investitie'!F86 +'3A-Proiectii_fin_investitie'!F89) /1.19 "
    ProiectiiFinaleInvestitie['F70'] = "=('3A-Proiectii_fin_investitie'!G86 +'3A-Proiectii_fin_investitie'!G89) /1.19 "

    ProiectiiFinaleInvestitie['C71']="=C70*0.19"
    ProiectiiFinaleInvestitie['D71']="=D70*0.19"
    ProiectiiFinaleInvestitie['E71']="=E70*0.19"
    ProiectiiFinaleInvestitie['F71']="=F70*0.19"

    ProiectiiFinaleInvestitie['C73']="= '3A-Proiectii_fin_investitie'!D82 /1.19"
    ProiectiiFinaleInvestitie['D73']="= '3A-Proiectii_fin_investitie'!E82 /1.19"
    ProiectiiFinaleInvestitie['E73']="= '3A-Proiectii_fin_investitie'!F82 /1.19"
    ProiectiiFinaleInvestitie['F73']="= '3A-Proiectii_fin_investitie'!G82 /1.19"

    ProiectiiFinaleInvestitie['C74'] = "=C73*0.19"
    ProiectiiFinaleInvestitie['D74'] = "=D73*0.19"
    ProiectiiFinaleInvestitie['E74'] = "=E73*0.19"
    ProiectiiFinaleInvestitie['F74'] = "=F73*0.19"

    
    
    ProiectiiFinaleInvestitie['C75'] = "='3A-Proiectii_fin_investitie'!D101"
    ProiectiiFinaleInvestitie['D75'] = "='3A-Proiectii_fin_investitie'!E101"
    ProiectiiFinaleInvestitie['E75'] = "='3A-Proiectii_fin_investitie'!F101"
    ProiectiiFinaleInvestitie['F75'] = "='3A-Proiectii_fin_investitie'!G101"

    ProiectiiFinaleInvestitie['C76'] = "=C75*2.25/100"
    ProiectiiFinaleInvestitie['C78'] ="='3A-Proiectii_fin_investitie'!D102 /1.19"
    ProiectiiFinaleInvestitie['D78'] ="='3A-Proiectii_fin_investitie'!E102 /1.19"
    ProiectiiFinaleInvestitie['E78'] ="='3A-Proiectii_fin_investitie'!F102 /1.19"
    ProiectiiFinaleInvestitie['F78'] ="='3A-Proiectii_fin_investitie'!G102 /1.19"

    ProiectiiFinaleInvestitie['C79'] ="=C78*0.19"
    ProiectiiFinaleInvestitie['C83'] = "='3A-Proiectii_fin_investitie'!D106"
    ProiectiiFinaleInvestitie['D83'] = "='3A-Proiectii_fin_investitie'!E106"
    ProiectiiFinaleInvestitie['E83'] = "='3A-Proiectii_fin_investitie'!F106"
    ProiectiiFinaleInvestitie['F83'] = "='3A-Proiectii_fin_investitie'!G106"

    ProiectiiFinaleInvestitie['C85'] ="='1B-ContPP'!C35"
    ProiectiiFinaleInvestitie['D85'] ="=C85*0.01 +C85"
    ProiectiiFinaleInvestitie['E85'] ="=D85*0.01 +D85"
    ProiectiiFinaleInvestitie['F85'] ="=E85*0.01 +E85"
    
    ProiectiiFinaleInvestitie['C83'] ="='3A-Proiectii_fin_investitie'!D106 - C85"
    ProiectiiFinaleInvestitie['D83'] ="='3A-Proiectii_fin_investitie'!D106 - D85"
    ProiectiiFinaleInvestitie['E83'] ="='3A-Proiectii_fin_investitie'!E106 - E85"
    ProiectiiFinaleInvestitie['F83'] ="='3A-Proiectii_fin_investitie'!F106 - F85"

    ProiectiiFinaleInvestitie['C97'] ="='1A-Bilant'!C28"
    
    ProiectiiFinaleInvestitie['C109'] = "=C53"
    ProiectiiFinaleInvestitie['D109'] = "=D53"
    ProiectiiFinaleInvestitie['E109'] = "=E53"
    ProiectiiFinaleInvestitie['F109'] = "=F53"

    ProiectiiFinaleInvestitie['C89'] = "=C39 + C42 +C45 + + C51+C54 + C48 - C28*19/119 - C65-C68-C71-C74-C79"
    ProiectiiFinaleInvestitie['D89'] = "=D39 + D42 +D45  + D51+D54 + D48 - D28*19/119 - D65-D68-D71-D74-D79"
    ProiectiiFinaleInvestitie['E89'] = "=E39 + E42 +E45  + E51+E54 + E48 - E28*19/119 - E65-E68-E71-E74-E79"
    ProiectiiFinaleInvestitie['F89'] = "=F39 + F42 +F45  + F51+F54 + F48 - F28*19/119 - F65-F68-F71-F74-F79"


    
    
    # ProiectiiFinaleInvestitie['C89'] = "=(C39 + C42 +C45 +C48 + C51+C54) - (C28*19/119) + (C65+C68+C71+C74+C79)"
    # if ProiectiiFinaleInvestitie['C89'].value >0:
    #     #ramane acolo
    # else:
    #     ProiectiiFinaleInvestitie['90'] = "=C89"
    
    ########################################################################
    
    ProiectiiRentabilitate['F9'] = 0
    
    #Modify the columns from ProiectiiFinInvestitie based on Bilant and PP
    
    """
    INCASARI DIN ACTIVITATEA DE EXPLOATARE (fara investitie)
    """
    ###
    #PretUnitarMarfaOrig = ProiectiiFinInvestitie['D17'].value 
    #ProiectiiInvestitieSim['D17'] = PretUnitarMarfaOrig
    ProiectiiInvestitieSim['D17'] ="='1B-ContPP'!C6*1.19*0.02 + '1B-ContPP'!C6*1.19"
    ProiectiiInvestitieSim['E17']= '=D17*0.01 + D17'#pret unitar (marfa)
    ProiectiiInvestitieSim['F17'] ='=D17*0.01 + E17'#pret unitar (marfa)
    ProiectiiInvestitieSim['G17'] ='=D17*0.01 +F17' #pret unitar (marfa)
    
    ProiectiiInvestitieSim['D15'] = '=D16*D17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['E15'] = '=E16*E17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['F15'] = '=F16*F17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['G15'] = '=G16*G17'#Venituri din vanzari marfuri
    ProiectiiInvestitieSim['C15'] = '=SUM(D15:G15)' #Total
    
    ProiectiiInvestitieSim['D18'] = '=D9+D12+D15'
    ProiectiiInvestitieSim['E18'] = '=E9+E12+E15'
    ProiectiiInvestitieSim['F18'] = '=F9+F12+F15'
    ProiectiiInvestitieSim['G18'] = '=G9+G12+G15'
    ProiectiiInvestitieSim['C18'] = '=SUM(D18:G18)'#Total incasari 
    
    ProiectiiInvestitieSim['D16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['E16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['F16'] =1 #cantitate marfuri
    ProiectiiInvestitieSim['G16'] =1 #cantitate marfuri
    ###
    
    """
    Cheltuieli de exploatare
    """
    ### 5.Cheltuieli cu materiile prime si cu materialele consumabile
    #PretUnitarMateriiPrimeOrig = ProiectiiFinInvestitie['D23'].value
    ProiectiiInvestitieSim['D23'] = "=('1B-ContPP'!C14 +'1B-ContPP'!C15)*1.19*0.01 + ('1B-ContPP'!C14 +'1B-ContPP'!C15)*1.19"
    ProiectiiInvestitieSim['E23'] = '=D23*0.005 + D23'
    ProiectiiInvestitieSim['F23'] = '=D23*0.005 +E23'
    ProiectiiInvestitieSim['G23'] = '=D23*0.005 +F23'
    
    ProiectiiInvestitieSim['D22'] = 1
    ProiectiiInvestitieSim['E22'] = 1
    ProiectiiInvestitieSim['F22'] = 1
    ProiectiiInvestitieSim['G22'] = 1
    
    ProiectiiInvestitieSim['D24'] = 1
    ProiectiiInvestitieSim['E24'] = 1
    ProiectiiInvestitieSim['F24'] = 1
    ProiectiiInvestitieSim['G24'] = 1
    
    #PretUnitarMaterialeConsumabile = ProiectiiFinInvestitie['D25'].value
    ProiectiiInvestitieSim['D25'] =0 ##Din C15
    ProiectiiInvestitieSim['E25'] ='=D25*1.05'
    ProiectiiInvestitieSim['F25'] ='=E25*1.05'
    ProiectiiInvestitieSim['G25'] ='=F25*1.05'
    
    ProiectiiInvestitieSim['D21'] ='=SUM(D22*D23)+SUM(D24*D25)'
    ProiectiiInvestitieSim['E21'] ='=SUM(E22*E23)+SUM(E24*E25)'
    ProiectiiInvestitieSim['F21'] ='=SUM(F22*F23)+SUM(F24*F25)'
    ProiectiiInvestitieSim['G21'] ='=SUM(G22*G23)+SUM(G24*G25)'
    ProiectiiInvestitieSim['C21'] = '=SUM(D21:G21)'
    
    ### 6.Cheltuieli privind marfurile 
    #PretUnitarMarfuri = ProiectiiFinInvestitie['D28'].value
    ProiectiiInvestitieSim['D28'] = "='1B-ContPP'!C17*1.19*0.005 + '1B-ContPP'!C17*1.19"
    ProiectiiInvestitieSim['E28'] = '=D28*0.005 + D28'
    ProiectiiInvestitieSim['F28'] = '=D28*0.005+E28'
    ProiectiiInvestitieSim['G28'] = '=D28*0.005 +F28'
    
    ProiectiiInvestitieSim['D27'] = 1
    ProiectiiInvestitieSim['E27'] = 1
    ProiectiiInvestitieSim['F27'] = 1
    ProiectiiInvestitieSim['G27'] = 1
    
    ProiectiiInvestitieSim['D26'] = '=D27*D28'
    ProiectiiInvestitieSim['E26'] = '=E27*E28'
    ProiectiiInvestitieSim['F26'] = '=F27*F28'
    ProiectiiInvestitieSim['G26'] = '=G27*G28'
    
    
    ### 7.Alte cheltuieli materiale (inclusiv cheltuieli cu prestatii externe)
    #AlteCheltuieliMateriale = ProiectiiFinInvestitie['D29'].value
    ProiectiiInvestitieSim['D29'] = 0 #C22
    ProiectiiInvestitieSim['E29'] = '=D29*1.01'
    ProiectiiInvestitieSim['F29'] = '=E29*1.01'
    ProiectiiInvestitieSim['G29'] = '=F29*1.01'
    ProiectiiInvestitieSim['C29']='=SUM(D29:G29)' #Total
    
    #D29 ramane pe 0 checked
    #D44 = C19*2.25% se modifica in functie de cheltuielile cu personalul
    #D46 = C22*..
    #D25 raamne pe 0 checked
    
    
    
    ### 8.Cheltuieli cu energia 
    ProiectiiInvestitieSim['D31'] = 1
    ProiectiiInvestitieSim['E31'] = 1
    ProiectiiInvestitieSim['F31'] = 1
    ProiectiiInvestitieSim['G31'] = 1
    
    #TarifFurnizareUnitar = ProiectiiFinInvestitie['D32'].value
    ProiectiiInvestitieSim['D32'] = "='1B-ContPP'!C16*1.19*0.01 + '1B-ContPP'!C16*1.19"
    ProiectiiInvestitieSim['E32']='=D32*0.01 + D32'
    ProiectiiInvestitieSim['F32']='=D32*0.01 + E32'
    ProiectiiInvestitieSim['G32']='=D32*0.01 + F32'
    
    ProiectiiInvestitieSim['D30'] = '=D31*D32'
    ProiectiiInvestitieSim['E30'] = '=E31*E32'
    ProiectiiInvestitieSim['F30'] = '=F31*F32'
    ProiectiiInvestitieSim['G30'] = '=G31*G32'
    
    ProiectiiInvestitieSim['C30'] ='=SUM(D30:G30)' #Total 
    
    ### 9.Cheltuieli cu apa
    
    ### 10.Alte cheltuieli (utilitati)
    
    
    ### Total Cheltuieli Materiale
    ProiectiiInvestitieSim['D39']='=D21+D26+D29+D30+D33+D36'
    ProiectiiInvestitieSim['E39']='=E21+E26+E29+E30+E33+E36'
    ProiectiiInvestitieSim['F39']='=F21+F26+F29+F30+F33+F36'
    ProiectiiInvestitieSim['G39']='=G21+G26+G29+G30+G33+G36'
    
    ProiectiiInvestitieSim['C39']= '=SUM(D39:G39)'
    
    
    ### 11.Cheltuieli cu personalul angajat
    ProiectiiInvestitieSim['D41']=1
    ProiectiiInvestitieSim['E41']=1
    ProiectiiInvestitieSim['F41']=1
    ProiectiiInvestitieSim['G41']=1
    
    #SalariuDeBaza =ProiectiiFinInvestitie['D42'].value
    ProiectiiInvestitieSim['D42'] = "='1B-ContPP'!C19"
    ProiectiiInvestitieSim['E42']='=D42*1'
    ProiectiiInvestitieSim['F42']='=E42*1'
    ProiectiiInvestitieSim['G42']='=F42*1'
    
    ProiectiiInvestitieSim['D43']=1
    ProiectiiInvestitieSim['E43']=1
    ProiectiiInvestitieSim['F43']=1
    ProiectiiInvestitieSim['G43']=1
    
    ProiectiiInvestitieSim['D40']='=(D41*D42)*D43'
    ProiectiiInvestitieSim['E40']='=(E41*E42)*E43'
    ProiectiiInvestitieSim['F40']='=(F41*F42)*F43'
    ProiectiiInvestitieSim['G40']='=(G41*G42)*G43'
    
    ProiectiiInvestitieSim['C40']='=SUM(D40:G40)' #Total
    
    ### 12.Cheltuieli cu asigurarile si protectia sociala 
    ProiectiiInvestitieSim['D44']="='1B-ContPP'!C19*0.0225"
    ProiectiiInvestitieSim['E44']= '=D44*1'
    ProiectiiInvestitieSim['F44']='=E44*1'
    ProiectiiInvestitieSim['G44']='=F44*1'
    ProiectiiInvestitieSim['C44']= '=SUM(D44:G44)' #Total
    
    ### Cheltuieli de personal 
    ProiectiiInvestitieSim['D45'] = '=D40+D44'
    ProiectiiInvestitieSim['E45'] = '=E40+E44'
    ProiectiiInvestitieSim['F45'] = '=F40+F44'
    ProiectiiInvestitieSim['G45'] = '=G40+G44'
    ProiectiiInvestitieSim['C45']= '=SUM(D45:G45)' #Total
    
    
    ### 13.Alte cheltuieli de exploatare (prestatii externe, alte impozite, taxe si varsaminte asimilate, alte cheltuieli), din care:
    #Tarif = ProiectiiFinInvestitie['D49'].value
    ProiectiiInvestitieSim['D49'] = 0
    ProiectiiInvestitieSim['E49'] = '=D49*1.01'
    ProiectiiInvestitieSim['F49'] = '=E49*1.01'
    ProiectiiInvestitieSim['G49'] = '=F49*1.01'
    
    ProiectiiInvestitieSim['D48'] = 1
    ProiectiiInvestitieSim['E48'] = 1
    ProiectiiInvestitieSim['F48'] = 1
    ProiectiiInvestitieSim['G48'] = 1
    
    ProiectiiInvestitieSim['D47'] = '=D48*D49'
    ProiectiiInvestitieSim['E47'] = '=E48*E49'
    ProiectiiInvestitieSim['F47'] = '=F48*F49'
    ProiectiiInvestitieSim['G47'] = '=G48*G49'
    
    ProiectiiInvestitieSim['C47']='=SUM(D47:G47)' #Total
    
    #Altele = ProiectiiFinInvestitie['D46'].value
   # AlteleModificat = Altele.replace('D22', 'C22') #corectie
    ProiectiiInvestitieSim['D46'] = "='1B-ContPP'!C22*1.19*0.01 + '1B-ContPP'!C22*1.19"
    ProiectiiInvestitieSim['E46'] ='=D46*0.005 + D46'
    ProiectiiInvestitieSim['F46'] ='=D46*0.005+E46'
    ProiectiiInvestitieSim['G46'] ='=D46*0.005+F46'
    
    ProiectiiInvestitieSim['C46']='=SUM(D46:G46)' #Total
    
    
    
    
    ### 14.Cheltuieli financiare 
    #CheltuieliFinanciare = ProiectiiFinInvestitie['D50'].value
    ProiectiiInvestitieSim['D50'] = "='1B-ContPP'!C36*0.01 +'1B-ContPP'!C36  "
    ProiectiiInvestitieSim['E50']='=D50*0.005 + D50'
    ProiectiiInvestitieSim['F50']='=D50*0.005 + E50'
    ProiectiiInvestitieSim['G50']='=D50*0.005 + F50'
    ProiectiiInvestitieSim['C50']='=SUM(D50:G50)'
    
    
    #PLATI TVA FARA INVESTITIE
    ProiectiiInvestitieSim['D53'] = "=D18*0.19/1.19 - (D21+D26+D29+D30+D46)*0.19/1.19"
    ProiectiiInvestitieSim['E53'] = "=E18*0.19/1.19 - (E21+E26+E29+E30+E46)*0.19/1.19"
    ProiectiiInvestitieSim['F53'] = "=F18*0.19/1.19 - (F21+F26+F29+F30+F46)*0.19/1.19"
    ProiectiiInvestitieSim['G53'] = "=G18*0.19/1.19 - (G21+G26+G29+G30+G46)*0.19/1.19"

    
    # Disponibil de numerar la inceputul perioadei 
    Numerar = "='1A-Bilant'!C28"
    ProiectiiInvestitieSim['D58'] = Numerar
    
    ### INCASARI DIN ACTIVITATEA DE EXPLOATARE  (cu adoptarea investitiei)
    
    ### 2.Venituri din prestari servicii
    #Venit = ProiectiiFinInvestitie['D70'].value
    ProiectiiInvestitieSim['D70'] = 0
    ProiectiiInvestitieSim['E70'] = 0
    ProiectiiInvestitieSim['F70'] = '=E70*1.02'
    ProiectiiInvestitieSim['G70'] = '=F70*1.03'

    ProiectiiInvestitieSim['D69'] = 1
    ProiectiiInvestitieSim['E69'] = 1
    ProiectiiInvestitieSim['F69'] = 1
    ProiectiiInvestitieSim['G69'] = 1
    
    ProiectiiInvestitieSim['D68']='=D69*D70'
    ProiectiiInvestitieSim['E68']='=E69*E70'
    ProiectiiInvestitieSim['F68']='=F69*F70'
    ProiectiiInvestitieSim['G68']='=G69*G70'
    
    ProiectiiInvestitieSim['C68']='=SUM(D68:G68)' #Total
    
    ### 3.Venituri din vanzari marfuri
    ProiectiiInvestitieSim['D73'] = '=D17'
    ProiectiiInvestitieSim['E73'] = '=D73*0.02 + D73'
    ProiectiiInvestitieSim['F73'] = '=D73*0.02 + E73'
    ProiectiiInvestitieSim['G73'] = '=D73*0.02 + F73'

    ProiectiiInvestitieSim['D72'] =1
    ProiectiiInvestitieSim['E72'] =1
    ProiectiiInvestitieSim['F72'] =1
    ProiectiiInvestitieSim['G72'] =1
    
    ProiectiiInvestitieSim['D71'] = '=D72*D73'
    ProiectiiInvestitieSim['E71'] = '=E72*E73'
    ProiectiiInvestitieSim['F71'] = '=F72*F73'
    ProiectiiInvestitieSim['G71'] = '=G72*G73'
    
    ProiectiiInvestitieSim['C71']= '=SUM(D71:G71)' #Total
    
    ProiectiiInvestitieSim['D74']= '=D65+D68+D71'
    ProiectiiInvestitieSim['E74']= '=E65+E68+E71'
    ProiectiiInvestitieSim['F74']= '=F65+F68+F71'
    ProiectiiInvestitieSim['G74']= '=G65+G68+G71'

    ProiectiiInvestitieSim['C74']= '=SUM(D74:G74)' #Total
    
    ### 5.Cheltuieli cu materiile prime si cu materialele consumabile
    
    ProiectiiInvestitieSim['D79'] ='=D23'   #Egalam cu precedenta din anul de implementare(Fara investitii)
    ProiectiiInvestitieSim['E79'] = '=D79*0.002 + D79'
    ProiectiiInvestitieSim['F79'] = '=D79*0.002 +E79'
    ProiectiiInvestitieSim['G79'] = '=D79*0.002 +F79'

    ProiectiiInvestitieSim['D78'] = 1
    ProiectiiInvestitieSim['E78'] = 1
    ProiectiiInvestitieSim['F78'] = 1
    ProiectiiInvestitieSim['G78'] = 1
    
    
    ProiectiiInvestitieSim['D77'] = '=SUM(D78*D79)+SUM(D80*D81)'
    ProiectiiInvestitieSim['E77'] = '=SUM(E78*E79)+SUM(E80*E81)'
    ProiectiiInvestitieSim['F77'] = '=SUM(F78*F79)+SUM(F80*F81)'
    ProiectiiInvestitieSim['G77'] = '=SUM(G78*G79)+SUM(G80*G81)'

    ProiectiiInvestitieSim['C77']= '=SUM(D77:G77)' #Total
    
    
    
    
    ### 6.Cheltuieli privind marfurile 
    
    ProiectiiInvestitieSim['D84'] = '=D28'
    ProiectiiInvestitieSim['E84'] = '=D84*0.005 + D84'
    ProiectiiInvestitieSim['F84'] = '=D84*0.005+E84'
    ProiectiiInvestitieSim['G84'] = '=D84*0.005+F84'
    
    ProiectiiInvestitieSim['D83'] = 1
    ProiectiiInvestitieSim['E83'] = 1
    ProiectiiInvestitieSim['F83'] = 1
    ProiectiiInvestitieSim['G83'] = 1
    
    ProiectiiInvestitieSim['D82'] = '=D83*D84'
    ProiectiiInvestitieSim['E82'] = '=E83*E84'
    ProiectiiInvestitieSim['F82'] = '=F83*F84'
    ProiectiiInvestitieSim['G82'] = '=G83*G84'
    
    ProiectiiInvestitieSim['C82'] = '=SUM(D82:G82)' #Total
    
    
    ### 7.Alte cheltuieli materiale (inclusiv cheltuieli cu prestatii externe)
    
    #Alte cheltuieli materiale  fara invvestitie == Alte cheltuieli investitie 
    ProiectiiInvestitieSim['D85'] = "=D29 "
    ProiectiiInvestitieSim['E85'] ='=D85*0.002 + D85'
    ProiectiiInvestitieSim['F85'] ='=D85*0.002 + E85'
    ProiectiiInvestitieSim['G85'] ='=D85*0.002 +F85'
    
    ProiectiiInvestitieSim['C85']='=SUM(D85:G85)' #Total
    
    ### 8.Cheltuieli cu energia +apa
    
    ProiectiiInvestitieSim['D88'] = '=D32'
    ProiectiiInvestitieSim['E88'] = '=D88*0.002 +D88'
    ProiectiiInvestitieSim['F88'] = '=D88*0.002 + E88'
    ProiectiiInvestitieSim['G88'] = '=D88*0.002 +F88'
    
    ProiectiiInvestitieSim['D87'] =1 
    ProiectiiInvestitieSim['E87'] =1 
    ProiectiiInvestitieSim['F87'] =1 
    ProiectiiInvestitieSim['G87'] =1 

    

    
    ### 9.Cheltuieli apa
    
    
    ### 10.Alte cheltuieli
    
    
    ### Total Cheltuieli Materiale
    ProiectiiInvestitieSim['D95']='=D77+D82+D85+D86+D89+D92'
    ProiectiiInvestitieSim['E95']='=E77+E82+E85+E86+E89+E92'
    ProiectiiInvestitieSim['F95']='=F77+F82+F85+F86+F89+F92'
    ProiectiiInvestitieSim['G95']='=G77+G82+G85+G86+G89+G92'
    
    ProiectiiInvestitieSim['C95']= '=SUM(D95:G95)'
    
    
    
    
    ### 11.Cheltuieli cu personal
    ProiectiiInvestitieSim['D99']=1
    ProiectiiInvestitieSim['E99']=1
    ProiectiiInvestitieSim['F99']=1
    ProiectiiInvestitieSim['G99']=1
    
    #Salar cu investitiii = Salar fara investitii pt primul an
    ProiectiiInvestitieSim['D98'] = "='1B-ContPP'!C19"
    ProiectiiInvestitieSim['E98']='=D98*0.010 + D98'
    ProiectiiInvestitieSim['F98']='=D98*0.010 + E98'
    ProiectiiInvestitieSim['G98']='=D98*0.010 + F98'
    
    ProiectiiInvestitieSim['D97'] =1
    ProiectiiInvestitieSim['E97'] =1
    ProiectiiInvestitieSim['F97']=1
    ProiectiiInvestitieSim['G97'] =1
    
    ProiectiiInvestitieSim['D96'] = '=SUM(D97*D98)*D99'
    ProiectiiInvestitieSim['E96'] = '=SUM(E97*E98)*E99'
    ProiectiiInvestitieSim['F96'] = '=SUM(F97*F98)*F99'
    ProiectiiInvestitieSim['G96'] = '=SUM(G97*G98)*G99'

    ProiectiiInvestitieSim['C96'] ='=SUM(D96:G96)'

    # 12.Cheltuieli cu asigurarile si protectia sociala 
    ProiectiiInvestitieSim['D100'] = '=D96*2.25%'
    ProiectiiInvestitieSim['E100'] = '=E96*2.25%'
    ProiectiiInvestitieSim['F100'] = '=F96*2.25%'
    ProiectiiInvestitieSim['G100'] = '=G96*2.25%'
    
    ProiectiiInvestitieSim['D101'] = '=D100+D96'
    ProiectiiInvestitieSim['E101'] = '=E100+E96'
    ProiectiiInvestitieSim['F101'] = '=F100+F96'
    ProiectiiInvestitieSim['G101'] = '=G100+G96'

    ProiectiiInvestitieSim['C101'] ='=SUM(D101:G101)'

    # 13.Alte cheltuieli de exploatare (prestatii externe, alte impozite, taxe si varsaminte asimilate, alte cheltuieli), din care:
    #AlteCheltuieli = ProiectiiFinInvestitie['D102'].value
    ProiectiiInvestitieSim['D102'] = '=D46'
    ProiectiiInvestitieSim['E102']='=D102*0.002 + D102'
    ProiectiiInvestitieSim['F102'] ='=D102*0.002 + E102'
    ProiectiiInvestitieSim['G102'] ='=D102*0.002 +F102'
    
    ProiectiiInvestitieSim['C102'] = '=SUM(D102:G102)'

    
    ProiectiiInvestitieSim['D101'] ='=D100+D96'
    ProiectiiInvestitieSim['E101'] ='=E100+E96'
    ProiectiiInvestitieSim['F101'] = '=F100+F96'
    ProiectiiInvestitieSim['G101'] = '=G100+G96'

    ProiectiiInvestitieSim['C101'] = '=SUM(D101:G101)'
    
    
    #Tarif proiecti investitie = Tarif proiect fara investitie (*primul an )
    ProiectiiInvestitieSim['D105'] = ProiectiiInvestitieSim['D49'].value
    ProiectiiInvestitieSim['E105'] ='=D105*0.02 + D105'
    ProiectiiInvestitieSim['F105'] ='=D105*0.03 + E105'
    ProiectiiInvestitieSim['G105'] ='=D105*0.04 + F105'
    
    ProiectiiInvestitieSim['D104'] =1
    ProiectiiInvestitieSim['E104'] =1
    ProiectiiInvestitieSim['F104'] =1
    ProiectiiInvestitieSim['G104'] =1
    
    ProiectiiInvestitieSim['D103']='=D104*D105'
    ProiectiiInvestitieSim['E103']='=E104*E105'
    ProiectiiInvestitieSim['F103']='=F104*F105'
    ProiectiiInvestitieSim['G103']='=G104*G105'

    ProiectiiInvestitieSim['C103']='=SUM(D103:G103)'

    # 14.Cheltuieli financiare (Cheltuieli privind dobanzile la imprumuturile contractate pentru activitatea aferenta investitiei)
    ProiectiiInvestitieSim['D106'] =  "=D50"
    ProiectiiInvestitieSim['E106'] = '=D106*0.01 + D106'
    ProiectiiInvestitieSim['F106'] = '=D106*0.01 +E106'
    ProiectiiInvestitieSim['G106'] = '=D106*0.01 +F106'
    ProiectiiInvestitieSim['C106']='=SUM(D106:G106)'
    
    
    ProiectiiInvestitieSim['D111'] = '=(D74/1.19) *(1/100)'
    ProiectiiInvestitieSim['E111'] = '=(E74/1.19)*(1/100)'
    ProiectiiInvestitieSim['F111'] = '=(F74/1.19)*(1/100)'
    ProiectiiInvestitieSim['G111'] = '=(G74/1.19)*(1/100)'
    
    
    ProiectiiFinaleInvestitie['C89'] = "= C39 + C42 +C45 +C48 + C51+C54 - C28*19/119 - C65-C68-C71-C74-C79"

    #PLATI TVA CU INVESTITIE
    ProiectiiInvestitieSim['D109'] ='=D74*0.19/1.19 -(D79+D81+D84+D85+D88+D91+D94+D102)*0.19/1.19'
    ProiectiiInvestitieSim['E109'] ='=E74*0.19/1.19 -(E79+E81+E84+E85+E88+E91+E94+E102)*0.19/1.19'
    ProiectiiInvestitieSim['F109'] ='=F74*0.19/1.19 -(F79+F81+F84+F85+F88+F91+F94+F102)*0.19/1.19'
    ProiectiiInvestitieSim['G109'] ='=G74*0.19/1.19 -(G79+G81+G84+F85+G88+G91+G94+G102)*0.19/1.19'
    
    
    ProiectiiInvestitieSim['D137'] = "='1A-Bilant'!C28"
    ProiectiiInvestitieSim['E137'] = "=D138"
    ProiectiiInvestitieSim['F137'] = "=E138"
    ProiectiiInvestitieSim['G137'] = "=F138"
    
    ProiectiiInvestitieSim['D138'] = '=D136 + D137'
    ProiectiiInvestitieSim['E138'] = '=E136 + E137'
    ProiectiiInvestitieSim['F138'] = '=F136 + F137'
    ProiectiiInvestitieSim['G138'] = '=G136 + G137'
 
    
    pathSave= path.split("\\")[-1].split('.')[0] + "completed.xlsx"
    
    pathSaveFinal= path.split("\\")[-1].split('.')[0] + "FINAL.xlsx"

    pathSave2=path.split("\\")[-1]
    
    pathSave2=path.replace(pathSave2, "")
    
    finalPath = os.path.join(pathSave2, pathSave)
    
    workbookSim.save(finalPath)
    
    
    workbook=xw.App(visible=False)
    workbook = xw.Book( r'C:\Users\alexandru.vesa\Desktop\Research\New_Personal_Program_DL_Programming\GitMy\my-work\FinanciaApplication\Test\proteincompleted.xlsx' )
    
    app=xw.apps.active

    
    workbookSim2=load_workbook(filename =finalPath)
    ProiectiiFinaleInvestitie =  workbookSim2['4-Proiectii_fin_intreprindere']


    
    if workbook.sheets['4-Proiectii_fin_intreprindere'].range('C89').value > 0:
          ProiectiiFinaleInvestitie['C89'] = "=C39 + C42 +C45 +C48 + C51+C54 + C48 - C28*19/119 - C65-C68-C71-C74-C79"
          ProiectiiFinaleInvestitie['D89'] = "=D39 + D42 +D45 +D48 + D51+D54 + D48 - D28*19/119 - D65-D68-D71-D74-D79"
          ProiectiiFinaleInvestitie['E89'] = "=E39 + E42 +E45 +E48 + E51+E54 + E48 - E28*19/119 - E65-E68-E71-E74-E79"
          ProiectiiFinaleInvestitie['F89'] = "=F39 + F42 +F45 +F48 + F51+F54 + F48 - F28*19/119 - F65-F68-F71-F74-F79"

          #workbook.close()

          workbook.save(os.path.join(pathSave2, pathSaveFinal))
          workbook.close()

         
    else:
        ProiectiiFinaleInvestitie['C90'] = "=C89"
        ProiectiiFinaleInvestitie['C89'] = 0
        #workbook.close()

        workbook.save(os.path.join(pathSave2, pathSaveFinal))
        workbook.close()

    
    workbook.close()
    
    return finalPath
    
    
def set_arguments():
    # Set arguments
    ap = argparse.ArgumentParser(description='Fill 3B sheet')

    ap.add_argument("-p", "--path", required=False, type=str)

    args = vars(ap.parse_args())
    return args

args = set_arguments()
modificareTabel1(args["path"])

    