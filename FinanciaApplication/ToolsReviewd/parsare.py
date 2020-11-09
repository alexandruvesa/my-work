# -*- coding: utf-8 -*-
"""
Created on Thu Oct 29 18:20:34 2020

@author: alexandru.vesa
"""
import pandas as pd
from openpyxl.utils import FORMULAE  
from openpyxl import load_workbook , Workbook
import xlwings as xw
import itertools
import numpy as np
import time
import os
import pandas as pd
import random



 
def evalIRR(path):
    workbook=xw.App(visible=False)
    workbook = xw.Book( path )
    RentabilitateInvestitie = workbook.sheets['3B-Rentabilitate_investitie'].range('B19').value
       
    return RentabilitateInvestitie



def getValue(string,workbook):
    ProiectiiFinInvestitie = workbook['3A-Proiectii_fin_investitie']
    

    
def analizaIndicator(indicator,combinatie):

    if indicator is int:
        pass
    else:
        splitString = indicator.split("*")[-1].split('+')[0]
        newFormula = indicator.replace(splitString, str(combinatie),1)
                    
    return newFormula

def procenteVenit():
    w1=np.arange(0.060, 0.065, 0.001)
    #w1 = np.array([0.094])
    w2=np.arange(0.035,0.045, 0.002)
    w3=np.arange(0.03, 0.035, 0.002)
    weights = [w1,w2,w3]
    listWeights =[]

    for i in itertools.product(*weights):
        listWeights.append(i)
        
    suma = [sum(listWeights[i]) for i in range(len(listWeights))]
    suma =np.array(suma)
    listWeights = np.array(listWeights)
    listaNew = listWeights[suma <0.17]
    return listaNew

def procenteCheltuieli():
    w1=np.array([0.003])
    w2=np.arange(0.003, 0.004, 0.0001)
    w3=np.arange(0.003, 0.004,0.0001)
    weights = [w1,w2,w3]
    listWeights =[]
    for i in itertools.product(*weights):
        listWeights.append(i)
        
    suma = [sum(listWeights[i]) for i in range(len(listWeights))]
    suma =np.array(suma)
    listWeights = np.array(listWeights)
    listaNew = listWeights[suma <0.10]
    return listaNew
    

def calcul(path ):
    
    venitProcente = procenteVenit()
    cheltuieliProcente=procenteCheltuieli()
    
    listaCombinatii = list(itertools.product(venitProcente, cheltuieliProcente))
    
    workbookFormula = load_workbook(filename = path)
    ProiectiiFinInvestitie = workbookFormula['3A-Proiectii_fin_investitie']
    
    indicatorAn1 = ProiectiiFinInvestitie['E73'].value
    indicatorAn2=ProiectiiFinInvestitie['F73'].value
    indicatorAn3=ProiectiiFinInvestitie['G73'].value
    
    
    cheltuieliAn1 = ProiectiiFinInvestitie['E84'].value
    cheltuieliAn2 = ProiectiiFinInvestitie['F84'].value
    cheltuieliAn3 = ProiectiiFinInvestitie['G84'].value
    print('pe aici')

    
    #workbook=xw.App(visible=False)
    #workbook = xw.Book( pathOrig )
    #An0 = workbook.sheets['3A-Proiectii_fin_investitie'].range('D73').value
    
    #Flux2 = workbook.sheets['3B-Rentabilitate_investitie'].range('D15').value
    #Flux3=workbook.sheets['3B-Rentabilitate_investitie'].range('E15').value
    #Flux4=workbook.sheets['3B-Rentabilitate_investitie'].range('F15').value
    
    pathIRR = path.split("\\")[-1]
    
    pathIRR = path.replace(pathIRR, "") 
    
    pathIRR= os.path.join(pathIRR, "IRR.xlsx")
    
    random.shuffle(listaCombinatii)
    for i in range(len(listaCombinatii)):
        try:
            formula1 = analizaIndicator(indicatorAn1, listaCombinatii[i][0][0])
            formula2 = analizaIndicator(indicatorAn2,listaCombinatii[i][0][1] )
            formula3 =analizaIndicator(indicatorAn3,listaCombinatii[i][0][2] )
            formula4 =analizaIndicator(cheltuieliAn1,listaCombinatii[i][1][0] )
            formula5 =analizaIndicator(cheltuieliAn2,listaCombinatii[i][1][1] )
            formula6 =analizaIndicator(cheltuieliAn3,listaCombinatii[i][1][2] )
            
            #load_workbook(filename = 'combinationInFor.xlsx')
            ProiectiiFinInvestitie['E73']= formula1
            ProiectiiFinInvestitie['F73']= formula2
            ProiectiiFinInvestitie['G73']= formula3
            ProiectiiFinInvestitie['E84']= formula4
            ProiectiiFinInvestitie['F84']= formula5
            ProiectiiFinInvestitie['G84']= formula6
                    
            workbookFormula.save(pathIRR)
            #workbookFormula._archive_close()
            
            path = os.getcwd()
            #originalPath = os.path.join(path, 'Test3.xlsx')
            print('mama')
            if i %50==0:
                
                workbook=xw.App(visible=False)
                workbook = xw.Book( pathIRR )
                app=xw.apps.active
                Flux2 = workbook.sheets['3B-Rentabilitate_investitie'].range('D15').value
                Flux3=workbook.sheets['3B-Rentabilitate_investitie'].range('E15').value
                Flux4=workbook.sheets['3B-Rentabilitate_investitie'].range('F15').value
                print("Flux2 : ", Flux2)
                print("Flux3 : " , Flux3)
                print("Flux4 : " , Flux4)
                
                if (Flux2 <0) or (Flux3<0) or (Flux4<0)  :
                    #workbook.save()
                    print("sunt pe bucla asta")
                    workbook.close()
                    continue
                if ((Flux2<Flux3<Flux4) and (Flux2>0) and (Flux3>0) and (Flux4>0)):
                    RentabilitateInvestitie = workbook.sheets['3B-Rentabilitate_investitie'].range('B19').value
                    print("IRR : ",RentabilitateInvestitie*100 )
                    originalPathNew=pathIRR.replace('IRR.xlsx', 'IRR.{}.xlsx'.format(RentabilitateInvestitie*100))
                    workbook.save(originalPathNew)
                    workbook.close()
    
    
                    if (RentabilitateInvestitie*100<=10 and RentabilitateInvestitie*100>=0):
                        workbook.save(originalPathNew)
                        workbook.close()
                        
                    if (RentabilitateInvestitie*100<=3 and RentabilitateInvestitie*100>=1):
                        workbook.save(originalPathNew)
                        workbook.close()
    
                        break
                #workbook.save()
                workbook.close()
        except:
                continue
     
            
        print(i)

path  = r'C:\Users\alexandru.vesa\Desktop\Research\New_Personal_Program_DL_Programming\GitMy\my-work\FinanciaApplication\Test\proteincompleted.xlsx'
        
calcul(path)      
        
        
        






                






